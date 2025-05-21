from openpyxl import load_workbook

def receber(nome):
    retorna = 0
    planilha = load_workbook("Sabor_20-05-2025 11-06-27.xlsx")
    plan = planilha.active
    for i in range(1,plan.max_row):
        if plan.cell(row=i,column=2).value.upper() == nome.upper():
            id = plan.cell(row=i,column=1)
            retorna = 1

    if retorna == 0:
        return "Item não Encontrado"
    else:
        return id.value

desc = ""
while desc != "sair":
    try:
        desc = input("Informe o nome do sabor: ")

        if desc == "sair":
            break

        elif receber(desc) == "Item não Encontrado":
            print(receber(desc))
            
        else:
            print(f"S{receber(desc)}")
    except:
        print('Algum erro aconteceu')
