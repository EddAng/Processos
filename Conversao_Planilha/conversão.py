from openpyxl import Workbook, load_workbook

#Função para a nova planilha receber o nome e o preço dos produtos
def preco(planilha_origem,planilha_destino,column_desc,column_prec,column_subg,column_cod,column_ncm):
    for linha in range(1,planilha_origem.max_row + 1):
        celula_des_desc = planilha_destino.cell(row=linha,column=1)
        celula_or_desc = planilha_origem.cell(row=linha,column=column_desc)
        celula_des_desc.value = celula_or_desc.value
        celula_des_prec = planilha_destino.cell(row=linha,column=5)
        celula_or_prec = planilha_origem.cell(row=linha,column=column_prec)
        celula_des_prec.value = celula_or_prec.value
        celula_des_subg = planilha_destino.cell(row=linha,column=2)
        celula_or_subg = planilha_origem.cell(row=linha,column=column_subg)
        celula_des_subg.value = celula_or_subg.value
        celula_des_cod = planilha_destino.cell(row=linha,column=14)
        celula_or_cod = planilha_origem.cell(row=linha,column=column_cod)
        celula_des_cod.value = celula_or_cod.value
        celula_des_ncm = planilha_destino.cell(row=linha,column=6)
        celula_or_ncm = planilha_origem.cell(row=linha,column=column_ncm)
        celula_des_ncm.value = celula_or_ncm.value
    
#Função para receber os valores caso tenha a mesma descrição que no banco
def recebe(planilha_destino,planilha_banco,linha_destino,linha_banco):
    subcategoria_or = planilha_destino.cell(row=linha_destino,column=3)
    subcategoria_bdo = planilha_banco.cell(row=linha_banco,column=3)
    subcategoria_or.value = subcategoria_bdo.value
    
    controla_or = planilha_destino.cell(row=linha_destino,column=4)
    controla_bdo = planilha_banco.cell(row=linha_banco,column=4)
    controla_or.value = controla_bdo.value

    icms_or = planilha_destino.cell(row=linha_destino,column=7)
    icms_bdo = planilha_banco.cell(row=linha_banco,column=7)
    icms_or.value = icms_bdo.value
    
    cest_or = planilha_destino.cell(row=linha_destino,column=8)
    cest_bdo = planilha_banco.cell(row=linha_banco,column=8)
    cest_or.value = cest_bdo.value
    
    pis_or = planilha_destino.cell(row=linha_destino,column=9)
    pis_bdo = planilha_banco.cell(row=linha_banco,column=9)
    pis_or.value = pis_bdo.value
    
    cofins_or = planilha_destino.cell(row=linha_destino,column=10)
    cofins_bdo = planilha_banco.cell(row=linha_banco,column=10)
    cofins_or.value = cofins_bdo.value
    
    unidade_or = planilha_destino.cell(row=linha_destino,column=11)
    unidade_bdo = planilha_banco.cell(row=linha_banco,column=11)
    unidade_or.value = unidade_bdo.value
    
#Função para validar se ja existe o cadastro no banco de dados
#Se existir, pegar as informações para o produto
def banco(planilha_banco,planilha_destino):
    planilha_destino['C1'] = 'SubCategoria'
    planilha_destino['D1'] = 'Controla estoque'
    planilha_destino['E1'] = 'Preço'
    planilha_destino['F1'] = 'Ncm'
    planilha_destino['G1'] = 'ICMS'
    planilha_destino['H1'] = 'Cest'
    planilha_destino['I1'] = 'Pis'
    planilha_destino['J1'] = 'Cofins'
    planilha_destino['K1'] = 'Unidade Medida'
    planilha_destino['L1'] = 'Local de Estoque Entrada'
    planilha_destino['M1'] = 'Local de Estoque Vendas'
    planilha_destino['N1'] = 'Código de Barras'
    
    for linha_dest in range(2,planilha_destino.max_row+1):
        for linha_banco in range(1,planilha_banco.max_row+1):
            if(planilha_destino.cell(row=linha_dest,column=1).value == planilha_banco.cell(row=linha_banco,column=1).value):
                recebe(planilha_destino,planilha_banco,linha_dest,linha_banco)


planilha_antiga = load_workbook('Produto.xlsx')
antigo = planilha_antiga.active
planilha_nova = load_workbook('Produtos Importar - ESSE.xlsx')
novo = planilha_nova.active
juncao = Workbook()
convercao = juncao.active

for numero_coluna in range(1,novo.max_column+1):
    if novo.cell(row=1,column=numero_coluna).value == 'NOMEREDUZIDO':
        desc = numero_coluna
        
    elif novo.cell(row=1,column=numero_coluna).value == 'VALOR':
        prec = numero_coluna
    elif novo.cell(row=1,column=numero_coluna).value == 'NOME':
        subgrupo = numero_coluna
    elif novo.cell(row=1,column=numero_coluna).value == 'CODIGOBARRA':
        codigo = numero_coluna
    elif novo.cell(row=1,column=numero_coluna).value == 'NCM':
        ncm = numero_coluna
    
preco(novo,convercao,desc,prec,subgrupo,codigo,ncm)
banco(antigo,convercao)

juncao.save('Panificadora_convertida.xlsx')