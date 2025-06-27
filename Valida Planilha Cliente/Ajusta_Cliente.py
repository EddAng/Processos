from openpyxl import load_workbook

class planilha:
    def __init__(self,nome_planilha:str,salvo,coluna_nome,coluna_telefone) -> None:
        self.planilha_nome = nome_planilha
        self.salvo = salvo
        self.ativa = salvo.active
        self.nome = coluna_nome
        self.tel = coluna_telefone
        
    def deleta_Duplicado(self,linha:int) -> None:
        for i in range((linha+1), (self.ativa.max_row+1)):
            if self.ativa.cell(row=linha,column=self.nome).value != "" or self.ativa.cell(row=linha,column=1).value !="":
                if self.ativa.cell(row=linha,column=self.nome).value == self.ativa.cell(row=i,column=self.nome).value and self.ativa.cell(row=linha,column=self.tel).value == self.ativa.cell(row=i,column=self.tel).value:
                    print(self.ativa.cell(row=i,column=self.tel).value)
                    self.ativa.delete_rows(i)

                elif self.ativa.cell(row=linha,column=self.tel).value == self.ativa.cell(row=i,column=self.tel).value:
                    self.ativa.cell(row=i,column=self.tel).value = ""

    def salvar_Planilha(self,nome) -> None:
        if nome == "":
            self.salvo.save(f"{self.planilha_nome}_copia.xlsx")
        else:
            self.salvo.save(f"{nome}.xlsx")
def valida_Planilha(valor:str,planilhas:list) -> bool:
    try:
        if len(planilhas) > 0:
            for i in range(0, len(planilhas)):
                if valor == planilhas[i]:
                    print("Planilha ja adicionada informe outra")
                    return False
        load_workbook(f"{valor}.xlsx")
        return True
    except Exception as e:
        print(e)
        return False
         
def ativar(nome_planilha:str,ativo,nome:str,telefone:str) -> None:
    for i in range (1,ativo.active.max_column+1):
        if ativo.active.cell(row=1,column=i).value == nome:
            column_nome = i
            
        if ativo.active.cell(row=1,column=i).value == telefone:
            column_tel = i
    return(planilha(salvo=ativo,coluna_nome=column_nome,coluna_telefone=column_tel,nome_planilha=nome_planilha))

def juntar(envia:classmethod,recebe:classmethod,linha:int) -> None:

    for i in range (1, recebe.ativa.max_row):
        if envia.ativa.cell(row=linha,column=envia.nome).value == recebe.ativa.cell(row=i,column=recebe.nome).value or envia.ativa.cell(row=linha,column=1).value == recebe.ativa.cell(row=i,column=1).value:
            return
        if recebe.ativa.cell(row=i,column=recebe.nome).value == "none" or recebe.ativa.cell(row=i,column=recebe.nome).value == None or (recebe.ativa.cell(row=i,column=1).value == "" and recebe.ativa.cell(row=i,column=recebe.nome).value == ""):
            for h in range(1,recebe.ativa.max_column):
                recebe.ativa.cell(row=i,column=h).value = envia.ativa.cell(row=linha,column=h).value
            for f in range(1,recebe.ativa.max_column):
                recebe.ativa.cell(row=(recebe.ativa.max_row+1),column=f).value = ""
            return
    for h in range(1,recebe.ativa.max_column):    
        recebe.ativa.cell(row=(recebe.ativa.max_row),column=h).value = envia.ativa.cell(row=linha,column=h).value
        for f in range(1,recebe.ativa.max_column):
            recebe.ativa.cell(row=(recebe.ativa.max_row+1),column=f).value = ""
    
nome_planilha = []
nome_plan = input("Informe o nome da planilha: ")
while valida_Planilha(valor=nome_plan,planilhas=nome_planilha) == False:
    nome_plan = input("Informe o nome da planilha: ")
nome_planilha.append(nome_plan)
    
nome = "Nome / Nome Fantasia"
telefone = "Tel Celular"

principal = ativar(nome_planilha=nome_plan,ativo=load_workbook(f"{nome_plan}.xlsx"),nome=nome,telefone=telefone)

for i in range (1, principal.ativa.max_row+1):
    principal.deleta_Duplicado(linha=i)

principal.salvar_Planilha("")
try:
    sair = ""
    while True:
        nome_plan = input("Informe o nome da planilha, Digite \"Sair\" para finalizar a aplicação: ")
        if nome_plan == "Sair":
            break
        
        while valida_Planilha(valor=nome_plan,planilhas=nome_planilha) == False:
            nome_plan = input("Informe o nome da planilha, Digite \"Sair\" para finalizar a aplicação: ")
            if nome_plan == "Sair":
                break
        if nome_plan == "Sair":
            break
        nome_planilha.append(nome_plan)
            
        nome = "Nome / Nome Fantasia"
        telefone = "Tel Celular"
         
        secundario = ativar(nome_planilha=nome_plan,ativo=load_workbook(f"{nome_plan}.xlsx"),nome=nome,telefone=telefone)

        for i in range (1, secundario.ativa.max_row+1):
            secundario.deleta_Duplicado(linha=i)

        secundario.salvar_Planilha("")
        
        for i in range (1, principal.ativa.max_row):
            if principal.ativa.cell(row=i,column=principal.nome).value == "none" or principal.ativa.cell(row=i,column=principal.nome).value == None:
                for f in range(1,principal.ativa.max_column):
                    principal.ativa.cell(row=(principal.ativa.max_row+1),column=f).value = ""
            
        for i in range (1, secundario.ativa.max_row):
            if secundario.ativa.cell(row=i,column=secundario.nome).value != "null":
                juntar(envia=secundario,recebe=principal,linha=i)
                
        principal.salvar_Planilha(input("Informe o nome da planilha: "))
except Exception as e:
    print(e)