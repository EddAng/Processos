from openpyxl import load_workbook


class Planilha():
    
    def __init__(self,nome,ativo):
        self.nome = nome
        self.ativo = ativo
        self.listaNumero = []
        self.listaNome = []
        self.listaRazaoSocial = []
        self.listaCPF = []
        self.listaLogradouro = []
        self.listaCEP = []
        self.listaCidade = []
        self.listaBairro = []
        self.listaEmail = []
        self.listaNumEnd = []
        
    def validaClientes(self):
        for i in range(2 ,(self.ativo.max_row)):
            passou = True
            if i==2:
                #telefone,razaoSocial,cpf,logradouro,cep,cidade,bairro,email
                adicionaDados(i,self.ativo,self.listaNome,self.listaNumero,self.listaRazaoSocial,
                    self.listaCPF,self.listaLogradouro,self.listaCEP,self.listaCidade,self.listaBairro,self.listaEmail,self.listaNumEnd)
                passou = False
            else:
                for x in range(0,len(self.listaNumero)):
                    if self.ativo.cell(row=i,column=2).value == self.listaNumero[x]:
                        passou = False
            if passou:
                adicionaDados(i,self.ativo,self.listaNome,self.listaNumero,self.listaRazaoSocial,
                    self.listaCPF,self.listaLogradouro,self.listaCEP,self.listaCidade,self.listaBairro,self.listaEmail,self.listaNumEnd)
     
    #Planilha[2]
    def comparaPlanilhas(self,planilhaA,planilhaB):
        metade = False
        quarto = False
        ultQuarto = False
        for i in range(0,(len(planilhaB.listaNumero)-1)):
            passou = True
            for x in range(0,(len(planilhaA.listaNumero)-1)):
                if planilhaA.listaNumero[x] == planilhaB.listaNumero[i]:
                    passou = False      
            if passou:        
                adiciona(i,planilhaB,self.listaNome,self.listaNumero,self.listaRazaoSocial,
                    self.listaCPF,self.listaLogradouro,self.listaCEP,self.listaCidade,self.listaBairro,self.listaEmail,self.listaNumEnd)
        
        
def adicionaDados(i,ativo,nome,telefone,razaoSocial,cpf,logradouro,cep,cidade,bairro,email,numEn):
    nome.append(ativo.cell(row=i,column=1).value)
    telefone.append(ativo.cell(row=i,column=2).value)
    razaoSocial.append(ativo.cell(row=i,column=3).value)
    cpf.append(ativo.cell(row=i,column=4).value)
    logradouro.append(ativo.cell(row=i,column=5).value)
    cep.append(ativo.cell(row=i,column=6).value)
    cidade.append(ativo.cell(row=i,column=7).value)
    bairro.append(ativo.cell(row=i,column=8).value)
    email.append(ativo.cell(row=i,column=9).value)
    numEn.append(ativo.cell(row=i,column=10).value)

def adiciona(i,ativo,nome,telefone,razaoSocial,cpf,logradouro,cep,cidade,bairro,email,numEn):
    nome.append(ativo.listaNome[i])
    telefone.append(ativo.listaNumero[i])
    razaoSocial.append(ativo.listaRazaoSocial[i])
    cpf.append(ativo.listaCPF[i])
    logradouro.append(ativo.listaLogradouro[i])
    cep.append(ativo.listaCEP[i])
    cidade.append(ativo.listaCidade[i])
    bairro.append(ativo.listaBairro[i])
    email.append(ativo.listaEmail[i])
    numEn.append(ativo.listaNumEnd[i])
    
    
        
def validaPlanilha(nome,planilhas):
    try:
        if len(planilhas)==0:
            planilhas.append(nome)
            return Planilha(load_workbook(nome),load_workbook(nome).active)
        elif nome not in planilhas:
            return Planilha(load_workbook(nome),load_workbook(nome).active)
        else:
            return False

    except Exception as e:
        print(e)
        
def imprimir(planilha,texto):
    for i in range(0,len(planilha.listaNumero)):
        texto.append(f"{planilha.listaNome[i]}-{planilha.listaNumero[i]}-{planilha.listaRazaoSocial[i]}-{planilha.listaCPF[i]}-{planilha.listaLogradouro[i]}-"+
            f"{planilha.listaCEP[i]}-{planilha.listaCidade[i]}-{planilha.listaBairro[i]}-{planilha.listaEmail[i]}-{planilha.listaNumEnd[i]}\n")
    input("Pressione Enter para sair")
    
nomeArquivo = input("Informe o nome do Arquivo a gerar: ")
try:
    arquivo = open(nomeArquivo+".txt", 'r+')
except Exception as e:
    print(e)
    arquivo = open(nomeArquivo+".txt", 'w+')
texto = arquivo.readlines()
planilhas = []
planilhasAtiva = []
planilhasAtiva.append(validaPlanilha((input("Informe o nome da Planilha de clientes Atual: ")+".xlsx"),planilhas))
planilhasAtiva.append(validaPlanilha((input("Informe o nome da Planilha de clientes Antiga: ")+".xlsx"),planilhas))
planilhasAtiva.append(Planilha("",""))
planilhasAtiva[0].validaClientes()
planilhasAtiva[1].validaClientes()
planilhasAtiva[2].comparaPlanilhas(planilhasAtiva[0],planilhasAtiva[1])
imprimir(planilhasAtiva[2],texto)
arquivo.writelines(texto)
arquivo.close()