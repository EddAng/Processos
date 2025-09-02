from openpyxl import load_workbook

class Planilha():
    def __init__(self,nome,ativo,tipo,desc,id):
        self.nome = nome
        self.ativo = ativo
        self.tipo = tipo
        self.produto_coluna = desc
        self.id_coluna = id
        
    def pegaCodigo(self,integracao):
        item = ""
        while item.upper() != "VOLTAR":
            passou = 0
            item = input(f"Informe o {self.tipo} desejado (\"Voltar\" para trocar planilha): ")
            for i in range(1,self.ativo.max_row):
                if item.upper() == self.ativo.cell(row=i,column=2).value.upper():
                    print(f"{item} - {codigo(integracao,self.tipo,self.ativo.cell(row=i,column=1).value)}")
                    passou = 1
            if passou == 0:
                palavra = item.split()
                for i in range(1,self.ativo.max_row):
                    c = 0
                    for b in palavra:
                        if b != "de" and "com" and "c/m" and "em":   
                            if b.upper() in self.ativo.cell(row=i,column=2).value.upper():
                                c = c + 1
                    if c >= 1:
                        print(f"{self.ativo.cell(row=i,column=2).value} - {codigo(integracao,self.tipo,self.ativo.cell(row=i,column=1).value)}")       
        
def validarPlanilha(nome,planilhas,tipo,integracao):
    try:
        if nome not in planilhas:
            planilha = load_workbook(nome+".xlsx") 
            ativo = planilha.active
            for i in range (1,ativo.max_column): 
                if ativo.cell(row=1,column=i).value.upper() == "DESCRIÇÃO":
                    descricao = i
                elif ativo.cell(row=1,column=i).value.upper() == "ID" and (integracao.upper() == "ANOTA AI" and tipo != "Pizza"):
                    id = i    
                elif ativo.cell(row=1,column=i).value.upper() == "ID" and integracao.upper() != "ANOTA AI":
                    id = i 
                elif ativo.cell(row=1,column=i).value.upper() == "PRODUTO ID:" and (integracao.upper() == "ANOTA AI" and tipo == "Pizza") :
                    id = i
            return Planilha(planilha,ativo,tipo,descricao,id)
        else:
            print("Planilha ja informada")
            return None
    except Exception as e:
        print(e)
        return None
     
def entra_Planilha(planilhas,planilha,integracao):
    tipo = None
    while not tipo:
        tipo = input(f"Informe o nome da planilha {planilha} \"Sair\" para não informar: ")
        if tipo == "Sair":
            return None
        else:
            tipo = validarPlanilha(tipo,planilhas,planilha,integracao)
            return tipo
        
def codigo(integracao,tipo, id):
    if integracao.upper() == "IFOOD":
        if tipo == "Sabor":
            return "S"+str(id)
        elif tipo == "Produto":
            return id
        elif tipo == "Adicional":
            return "AD"+str(id)
        elif tipo == "Pizza":
            return "PZ"+str(id)
    elif integracao.upper() == "ANOTA AI":    
        if tipo == "Sabor":
            return "S"+str(id)
        elif tipo == "Produto":
            return id
        elif tipo == "Adicional":
            return id
        elif tipo == "Pizza":
            return id
    
integracao = ""
while integracao.upper() != "IFOOD" and integracao.upper() != "ANOTA AI":
    integracao = input("Informe o tipo da integração (Ifood, Anota Ai): ")

planilhas = []

produto = None
produto = entra_Planilha(planilhas,"Produto",integracao=integracao)

pizza = None
pizza = entra_Planilha(planilhas,"Pizza",integracao=integracao)

adicional = None
adicional = entra_Planilha(planilhas,"Adicional",integracao=integracao)

sabor = None
sabor = entra_Planilha(planilhas,"Sabor",integracao=integracao)

tipo = ""

while tipo != "SAIR":
    tipo = input("Informe o tipo da planilha, (\"Sair\" para finalizar): ")
    if tipo == "Produto" and produto != None:
        produto.pegaCodigo(integracao)
    elif tipo == "Pizza" and pizza != None:
        pizza.pegaCodigo(integracao)
    elif tipo == "Sabor" and sabor != None:
        sabor.pegaCodigo(integracao)
    elif tipo == "Adicional" and adicional != None:
        adicional.pegaCodigo(integracao)
    elif tipo == "Sair":
        break