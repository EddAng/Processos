from openpyxl import load_workbook

def valida(planilha,numero):
    for numero_linha in range(numero+1,planilha.max_row+1):
        if planilha.cell(row=numero,column=9).value == planilha.cell(row=numero_linha,column=9).value and planilha.cell(row=numero_linha,column=9).value != "":
            planilha.cell(row=numero_linha,column=8).value = "Duplicado"
        

planilha = load_workbook("Cliente_06-05-2025 11-50-06.xlsx")
ativo = planilha.active

for i in range(1,ativo.max_row+1):
    valida(planilha=ativo,numero=i)
    
planilha.save("Cliente_06-05-2025 11-50-06 editado.xlsx")

#12,15