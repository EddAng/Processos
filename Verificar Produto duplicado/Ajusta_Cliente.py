from openpyxl import load_workbook

class planilha:
    def __init__(self,salvo:classmethod,coluna_nome,coluna_telefone) -> None:
        self.salvo = salvo
        self.ativa = salvo.active
        self.nome = coluna_nome
        self.tel = coluna_telefone
        
    def deleta_Duplicado(self,linha:int) -> None:
        for i in range((linha+1), (self.ativa.max_row+1)):
            if self.ativa.cell(row=linha,column=self.nome).value == self.ativa.cell(row=i,column=self.nome).value and self.ativa.cell(row=linha,column=self.tel).value == self.ativa.cell(row=i,column=self.tel).value:
                print(self.ativa.cell(row=i,column=self.tel).value)
                ativo.ativa.delete_rows(i)

            elif self.ativa.cell(row=linha,column=self.tel).value == self.ativa.cell(row=i,column=self.tel).value:
                self.ativa.cell(row=i,column=self.tel).value = ""

def ativar(ativo:classmethod,nome:str,telefone:str) -> classmethod:
    for i in range (1,load_workbook("Cliente_06-05-2025 11-50-06.xlsx").active.max_column+1):
        if ativo.cell(row=1,column=i).value == nome:
            column_nome = i
            
        if ativo.cell(row=1,column=i).value == telefone:
            column_tel = i
    return(planilha(salvo=load_workbook("Cliente_06-05-2025 11-50-06.xlsx"),coluna_nome=column_nome,coluna_telefone=column_tel))

nome = input("Informe qual o nome da Coluna Nome: ")
telefone = input("Informe o nome da coluna telefone: ")
ativo = ativar(ativo=load_workbook("Cliente_06-05-2025 11-50-06.xlsx").active,nome=nome,telefone=telefone)

for i in range (1, ativo.ativa.max_row+1):
    ativo.deleta_Duplicado(linha=i)
    
ativo.salvo.save("Teste.xlsx")